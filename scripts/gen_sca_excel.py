"""生成 gin-vuln-demo SCA 扫描结果记录表（12 场景版）"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 颜色常量 ──────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3864"   # 深蓝（表头背景）
C_FILL_BG    = "00B0F0"   # 亮蓝（预埋信息底色）
C_RESULT_BG  = "E2EFDA"   # 浅绿（待填写区底色）
C_SECTION_BG = "D6E4F7"   # 浅蓝（分组行）
C_ALT_ROW    = "F2F7FF"   # 交替行淡蓝

def fill(color):
    return PatternFill("solid", fgColor=color)

def border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def align_c(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def align_l(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def font_h(size=10, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=True, color=color)

def font_b(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def title_row(ws, text, cols, row=1, height=30, size=13):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(C_HEADER_BG)
    c.font = font_h(size)
    c.alignment = align_c()
    ws.row_dimensions[row].height = height

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1：CWE 应用层漏洞（SAST）
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "CWE 应用层漏洞"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

for i, w in enumerate([5,14,18,32,22,10,26,26,14,18,22,26], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

title_row(ws1, "gin-vuln-demo  ·  CWE 应用层漏洞扫描结果记录表", 12)
ws1.row_dimensions[2].height = 42

headers1 = [
    "#", "模块", "漏洞类型", "漏洞描述", "触发路由", "CWE",
    "预埋文件", "预埋触发点",
    "工具是否检出\n(Y/N)", "检出等级", "检出位置\n(文件:行号)", "备注"
]
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.fill = fill(C_FILL_BG) if col <= 8 else fill(C_RESULT_BG)
    c.font = font_h(color="000000")
    c.alignment = align_c(wrap=True)
    c.border = border()

cwe_rows = [
    (1, "player",   "SQL 注入",   "直接拼接用户输入构造 SQL，未使用参数化查询",     "GET /player/info",      "CWE-89",  "internal/handler/player.go",   "PlayerInfo → DB.Raw(sql+id)"),
    (2, "recharge", "XSS",        "用户输入 note 直接写入 HTML 响应，未 HTML 转义", "POST /recharge/submit", "CWE-79",  "internal/handler/recharge.go", "RechargeSubmit → fmt.Sprintf(html,note)"),
    (3, "bet",      "路径遍历",   "filename 未过滤 ../，可读取任意文件",             "GET /bet/record",       "CWE-22",  "internal/handler/bet.go",      "BetRecord → os.ReadFile(base+filename)"),
    (4, "item",     "SSRF",       "用户输入 url 直接传给 http.Get，可访问内网",     "GET /item/fetch",       "CWE-918", "internal/handler/item.go",     "ItemFetch → http.Get(url)"),
    (5, "withdraw", "命令注入",   "account 拼接到 shell 命令，exec.Command+sh -c",  "POST /withdraw/check",  "CWE-78",  "internal/handler/withdraw.go", "WithdrawCheck → exec.Command(sh,-c,cmd)"),
    (6, "save",     "XXE",        "解析 XML 未禁用外部实体，可读本地文件/SSRF",     "POST /save/profile",    "CWE-611", "internal/handler/save.go",     "SaveProfile → xml.Unmarshal(body)"),
]

for r, row in enumerate(cwe_rows, 3):
    ws1.row_dimensions[r].height = 34
    bg = "FFFFFF" if r % 2 == 1 else C_ALT_ROW
    for col, val in enumerate(row, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.border = border()
        if col <= 8:
            c.fill = fill(bg)
            c.font = font_b()
            c.alignment = align_c() if col in (1, 6) else align_l()
        else:
            c.fill = fill(C_RESULT_BG)
            c.font = font_b(color="999999")
            c.alignment = align_c(wrap=True)
            c.value = ""

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2：SCA 依赖场景（12 场景）
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("SCA 依赖场景")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

for i, w in enumerate([4, 16, 20, 36, 32, 26, 28, 14, 20, 30, 26], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

title_row(ws2, "gin-vuln-demo  ·  Go SCA 依赖场景扫描结果记录表（12 场景）", 11)
ws2.row_dimensions[2].height = 48

headers2 = [
    "#", "场景类型", "场景名称", "场景描述",
    "预埋包 / 版本", "CVE / 风险标识", "预埋文件",
    "工具是否检出\n(Y/N)", "检出等级\n(Critical/High/Med/Low/Info)",
    "检出内容摘要", "备注"
]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = fill(C_FILL_BG) if col <= 7 else fill(C_RESULT_BG)
    c.font = font_h(color="000000")
    c.alignment = align_c(wrap=True)
    c.border = border()

# 场景类型 → 行底色
type_color = {
    "漏洞":   "FFF2CC",
    "维护":   "FCE4D6",
    "作用域": "DDEBF7",
    "版本":   "EDE7F6",
    "完整性": "FDE9D9",
    "代码指纹": "F0F0F0",
    "工作区": "E8F5E9",
}

sca_rows = [
    # 漏洞类
    ("漏洞", "直接依赖 CVE",
     "yaml.Unmarshal 对超大输入解析耗时呈指数增长，可导致 CPU DoS",
     "gopkg.in/yaml.v2 v2.2.2",
     "CVE-2022-3064 (High)",
     "internal/helper/vuln_yaml.go"),
    ("漏洞", "直接依赖 CVE",
     "Compression 扩展处理 WebSocket 帧可导致内存耗尽 DoS",
     "github.com/gorilla/websocket v1.4.1",
     "CVE-2020-27813 (High)",
     "internal/helper/vuln_websocket.go"),
    ("漏洞", "直接依赖 CVE",
     "audience claim 校验不严格，可绕过 JWT token 验证",
     "github.com/dgrijalva/jwt-go v3.2.0+incompatible",
     "CVE-2020-26160 (High)",
     "internal/helper/vuln_jwt.go"),
    ("漏洞", "间接依赖 CVE",
     "ghodss/yaml 传递依赖 yaml.v2 v2.2.2，CVE 在间接依赖层",
     "github.com/ghodss/yaml v1.0.0\n→ gopkg.in/yaml.v2 v2.2.2",
     "CVE-2022-3064 (间接路径)",
     "internal/helper/vuln_ghodss.go"),
    # 维护类
    ("维护", "Archived 废弃包",
     "仓库已被作者 archived，无人维护，未来漏洞不会被修复",
     "github.com/russross/blackfriday v1.6.0",
     "无直接 CVE（维护终止风险）",
     "internal/helper/vuln_blackfriday.go"),
    # 作用域类
    ("作用域", "测试包流入生产",
     "Go 无 require-dev 机制，testify 被非 _test.go 文件 import，进入生产构建",
     "github.com/stretchr/testify v1.11.1",
     "无 CVE（作用域风险）",
     "internal/helper/debug_helper.go"),
    # 版本类
    ("版本", "Go EOL + toolchain",
     "go 1.21 进入维护末段；toolchain go1.21.0 含已知漏洞",
     "go.mod: go 1.21 + toolchain go1.21.0",
     "CVE-2023-39323 / CVE-2023-39325",
     "go.mod"),
    ("版本", "stdlib CVE",
     "net/http 不在 go.mod 中，stdlib CVE 由工具链版本决定，需特殊关联",
     "net/http（go1.21.0 工具链）",
     "CVE-2023-39325（HTTP/2 rapid reset）",
     "internal/helper/stdlib_use.go"),
    # 完整性类
    ("完整性", "replace 本地路径",
     "replace 指向本地目录，go.sum 不记录哈希，内容可任意修改",
     "github.com/gorilla/mux → ./packages/local-router",
     "供应链完整性风险",
     "go.mod replace 指令\npackages/local-router/"),
    # 工作区类
    ("工作区", "go.work workspace 覆盖",
     "go.work 优先级高于 go.mod；不感知 go.work 的工具对 yaml.v2 CVE 存在误报",
     "go.work: stub-yaml-patched 覆盖 yaml.v2 v2.2.2",
     "CVE-2022-3064（工具感知差异）",
     "go.work\npackages/stub-yaml-patched/"),
    # 代码指纹类
    ("代码指纹", "拷贝代码（Embedded OSS）",
     "直接拷贝 yaml.v2 源码片段并保留版权头，不经 go.mod，依赖图扫描无法发现",
     "yaml.v2 v2.2.2 decode.go 片段",
     "CVE-2022-3064（代码指纹路径）",
     "internal/helper/yaml_copy.go"),
    ("代码指纹", "go:embed 嵌入漏洞文件",
     "jQuery 1.8.3 编译进二进制，不在依赖图中；文件存在于源码目录，文件指纹可识别",
     "internal/helper/static/jquery-1.8.3.min.js\n（via //go:embed）",
     "CVE-2015-9251 / CVE-2019-11358",
     "internal/helper/static_embed.go"),
]

for r, row in enumerate(sca_rows, 3):
    ws2.row_dimensions[r].height = 48
    bg = type_color.get(row[0], "FFFFFF")
    # # 列
    c0 = ws2.cell(row=r, column=1, value=r - 2)
    c0.fill = fill(bg); c0.font = font_b(bold=True)
    c0.alignment = align_c(); c0.border = border()
    # 预埋列
    for col, val in enumerate(row, 2):
        c = ws2.cell(row=r, column=col, value=val)
        c.fill = fill(bg); c.font = font_b()
        c.alignment = align_l(); c.border = border()
    # 结果列（留空）
    for col in range(8, 12):
        c = ws2.cell(row=r, column=col)
        c.fill = fill(C_RESULT_BG); c.font = font_b(color="999999")
        c.alignment = align_c(wrap=True); c.border = border()
        c.value = ""

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3：工具能力对比
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("工具能力对比")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "B3"

ws3.column_dimensions["A"].width = 42
for col in "BCDEFG":
    ws3.column_dimensions[col].width = 20

title_row(ws3, "SCA 工具检出能力横向对比（每列填一个工具）", 7)
ws3.row_dimensions[2].height = 48

# 表头
c = ws3.cell(row=2, column=1, value="场景 / 检测项")
c.fill = fill(C_FILL_BG); c.font = font_h(color="000000")
c.alignment = align_c(wrap=True); c.border = border()

for col in range(2, 7):
    c = ws3.cell(row=2, column=col,
                 value=("最低所需\n扫描能力" if col == 6 else f"工具 {col-1}\n（填写工具名）"))
    c.fill = fill(C_SECTION_BG) if col == 6 else fill(C_RESULT_BG)
    c.font = font_h(color="000000")
    c.alignment = align_c(wrap=True); c.border = border()

compare_rows = [
    ("场景1:  yaml.v2 v2.2.2 / CVE-2022-3064（直接依赖）",         "go.mod 依赖图扫描"),
    ("场景2:  gorilla/websocket v1.4.1 / CVE-2020-27813（直接）",   "go.mod 依赖图扫描"),
    ("场景3:  dgrijalva/jwt-go v3.2.0 / CVE-2020-26160（直接）",    "go.mod 依赖图扫描"),
    ("场景4:  ghodss/yaml → yaml.v2 CVE-2022-3064（间接依赖）",     "传递依赖图分析"),
    ("场景5:  blackfriday v1.6.0 archived 废弃包",                   "仓库状态感知"),
    ("场景6:  testify 测试包流入生产（非 _test.go）",                "调用图 / import 分析"),
    ("场景7:  go 1.21 + toolchain go1.21.0 EOL / CVE",              "go 版本 CVE 关联"),
    ("场景8:  net/http stdlib CVE（工具链版本关联）",                "stdlib CVE 数据库"),
    ("场景9:  gorilla/mux replace 本地路径（供应链完整性）",         "replace 指令解析"),
    ("场景10: go.work 覆盖 yaml.v2 CVE（工具感知差异）",            "go.work 感知"),
    ("场景11: yaml.v2 拷贝代码片段（Embedded OSS）",                "代码片段指纹扫描"),
    ("场景12: jQuery 1.8.3 go:embed 嵌入漏洞文件",                  "文件内容指纹扫描"),
]

row_colors = ["FFFFFF", C_ALT_ROW]
for r, (name, level) in enumerate(compare_rows, 3):
    ws3.row_dimensions[r].height = 30
    bg = row_colors[r % 2]
    c1 = ws3.cell(row=r, column=1, value=name)
    c1.fill = fill(bg); c1.font = font_b()
    c1.alignment = align_l(wrap=False); c1.border = border()

    for col in range(2, 6):
        c = ws3.cell(row=r, column=col)
        c.fill = fill(C_RESULT_BG); c.font = font_b(color="999999")
        c.alignment = align_c(); c.border = border(); c.value = ""

    c6 = ws3.cell(row=r, column=6, value=level)
    c6.fill = fill(C_SECTION_BG); c6.font = font_b(bold=True)
    c6.alignment = align_c(wrap=True); c6.border = border()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4：说明
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("说明")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 68

title_row(ws4, "使用说明", 2, size=13)

desc = [
    ("工作表", ""),
    ("CWE 应用层漏洞", "6 个 CWE 漏洞模块的 SAST 扫描结果（SQL注入/XSS/路径遍历/SSRF/命令注入/XXE）"),
    ("SCA 依赖场景",   "12 类 Go SCA 预埋场景逐条记录扫描结果"),
    ("工具能力对比",   "多工具横向对比表，每列填一个工具名及其对各场景的检出情况"),
    ("", ""),
    ("填写约定", ""),
    ("工具是否检出", "Y = 检出  /  N = 未检出  /  P = 部分检出"),
    ("检出等级",     "按工具实际报告填写：Critical / High / Medium / Low / Info"),
    ("检出位置",     "填写工具报告的文件路径+行号，如 internal/helper/vuln_yaml.go:10"),
    ("", ""),
    ("底色说明", ""),
    ("蓝色底色", "预埋信息（只读参考）"),
    ("绿色底色", "待填写区（扫描结果）"),
    ("浅蓝底色", "能力层级参考（只读）"),
    ("", ""),
    ("场景说明", ""),
    ("场景 1-4",  "依赖 CVE：go.mod 中存在已知漏洞的直接/间接依赖包"),
    ("场景 5",    "废弃包：仓库已 archived，维护终止风险"),
    ("场景 6",    "作用域：测试包 testify 被非测试文件引入生产"),
    ("场景 7-8",  "版本：go 1.21 EOL + toolchain CVE + stdlib CVE"),
    ("场景 9",    "完整性：replace 本地路径，go.sum 无哈希校验"),
    ("场景 10",   "工作区：go.work 覆盖 go.mod，工具感知差异"),
    ("场景 11-12","代码指纹：拷贝代码片段 + go:embed 嵌入漏洞 JS 文件"),
]

for r, (k, v) in enumerate(desc, 2):
    ws4.row_dimensions[r].height = 22
    c1 = ws4.cell(row=r, column=1, value=k)
    c2 = ws4.cell(row=r, column=2, value=v)
    if v == "" and k:
        c1.fill = fill(C_SECTION_BG)
        c1.font = font_b(bold=True)
        ws4.merge_cells(f"A{r}:B{r}")
    else:
        c1.font = font_b(bold=bool(k))
        c2.font = font_b()
    c1.alignment = align_l(wrap=False)
    c2.alignment = align_l(wrap=True)
    c1.border = border(); c2.border = border()

# ── 保存 ──────────────────────────────────────────────────────────────────────
out_path = "security-audit/sca/sca-scan-results.xlsx"
wb.save(out_path)
print(f"Excel 生成成功：{out_path}")
